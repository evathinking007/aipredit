import warnings
import openpyxl
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
from datetime import timedelta
from datetime import datetime
import time
import requests
import logging


class AIPredict:
    def __init__(self):
        pass

    def clean_data(self, file_path):
        bitcoin_data = pd.read_excel(file_path)
        # bitcoin_data.drop(bitcoin_data.index, inplace=True)
        bitcoin_data = bitcoin_data.iloc[0:0]
        # 替换为您的实际表头
        bitcoin_data.columns = ['Date', 'Price', 'Volume', 'MA', 'RSI', 'Upper_BB', 'Lower_BB', 'Future_Price']
        # bitcoin_data['Future_Price'] = None
        bitcoin_data.to_excel(file_path, index=False)

    def caculate_data(self, file_path):
        df = pd.read_excel(file_path)
        ma_period = 20  # 移动平均的周期
        # if not pd.isna(df.at[1, "MA"]):
        #     return
        df['MA'] = df['Price'].rolling(window=ma_period).mean()

        # 计算相对强弱指标（RSI）
        rsi_period = 14  # RSI的周期
        price_diff = df['Price'].diff()
        gain = price_diff.where(price_diff > 0, 0)
        loss = -price_diff.where(price_diff < 0, 0)
        avg_gain = gain.rolling(window=rsi_period).mean()
        avg_loss = loss.rolling(window=rsi_period).mean()
        rs = avg_gain / avg_loss
        rsi = 100 - (100 / (1 + rs))
        df['RSI'] = rsi

        # 计算布林带的上限（Upper_BB）和下限（Lower_BB）
        bb_period = 20  # 布林带的周期
        std_dev = df['Price'].rolling(window=bb_period).std()
        df['Upper_BB'] = df['MA'] + (2 * std_dev)
        df['Lower_BB'] = df['MA'] - (2 * std_dev)

        # 清楚前20条没有数据的行
        df = df.iloc[20:]
        # df['Date'] = df['Date'].str.replace('UTC+0', '')
        df['Future_Price'] = None

        # 回写到Excel文件
        df.to_excel(file_path, index=False)

    def train_data(self, bitcoin_data):
        # 选择特征和目标变量
        # features = bitcoin_data[['Volume', 'MA', 'RSI', 'Upper_BB', 'Lower_BB']]

        features = bitcoin_data[['Volume', 'MA']]
        target = bitcoin_data['Price']

        # 划分数据集为训练集和测试集
        X_train, X_test, y_train, y_test = train_test_split(features, target, test_size=0.2, random_state=42)

        # 创建并拟合线性回归模型
        model = LinearRegression()

        model.fit(X_train, y_train)

        # 预测测试集的价格
        y_pred = model.predict(X_test)
        # logging.info(y_pred)
        # 评估模型性能
        mse = mean_squared_error(y_test, y_pred)
        r2 = r2_score(y_test, y_pred)

        logging.info(f"均方误差 (MSE): {mse}")
        logging.info(f"决定系数 (R^2): {r2}")
        warnings.filterwarnings("ignore", category=UserWarning)
        return model

    def predict_next(self, model, bitcoin_data):
        # 使用模型进行未来价格预测
        # bitcoin_data= bitcoin_data.iloc[20:]
        last_feature = bitcoin_data[['Volume', 'MA']].iloc[-1].values
        future_price = model.predict([last_feature])
        return round(future_price[0],2)

    def sort_date_value(self, file_path):
        bitcoin_data = pd.read_excel(file_path)
        bitcoin_data.sort_values(by='Date', ascending=True)
        bitcoin_data.to_excel(file_path, index=False)


class GateIO_Api:
    def __init__(self):
        pass

    def get_right_time(self, minutes_to_add):
        # 获取当前时间
        current_time = datetime.now()
        # 计算下一个整点时刻
        next_time = current_time + timedelta(minutes=minutes_to_add)
        next_time = next_time.replace(second=0)

        # 计算等待时间，直到下一个整点时刻
        time_to_wait = (next_time - current_time).total_seconds()
        logging.info(f"等待到下一个整点时刻（{next_time.strftime('%d/%m/%Y, %H:%M:%S')}）")
        time.sleep(time_to_wait)

    def get_current_price(self):
        host = "https://api.gateio.ws"
        prefix = "/api/v4"
        headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}

        url = '/spot/candlesticks'
        query_param = 'currency_pair=ETH_USDT&limit=1'
        r = requests.request('GET', host + prefix + url + "?" + query_param, headers=headers)
        data_k_price = r.json()[0][2]
        return data_k_price

    def get_current_data_api(self, file_path):
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        worksheet = workbook.active

        host = "https://api.gateio.ws"
        prefix = "/api/v4"
        headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}

        url = '/spot/candlesticks'
        query_param = 'currency_pair=ETH_USDT&limit=1000&interval=5m'

        r = requests.request('GET', host + prefix + url + "?" + query_param, headers=headers)
        data_rows_num = len(r.json())
        logging.info("===========get " + str(data_rows_num) + " k datas===========")
        worksheet.insert_rows(data_rows_num)
        for i in range(2, data_rows_num):
            # data_k_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            data_k_date = datetime.fromtimestamp(int(r.json()[i][0]))
            data_k_price = r.json()[i][2]
            data_k_volume = r.json()[i][6]
            # item = {
            #     'Date': [data_k_date],  # 示例时间
            #     'Price': [data_k_price],  # 示例价格
            #     'Volume': [data_k_volume],  # 示例成交量
            #     'MA': [None],
            #     'RSI': [None],
            #     'Upper_BB': [None],
            #     'Lower_BB': [None],
            #     'Future_Price': [None],
            # }
            # # # logging.info(new_data)
            # bitcoin_data.append(item, ignore_index=True)
            worksheet.cell(row=i, column=1, value=data_k_date)
            worksheet.cell(row=i, column=2, value=float(data_k_price))
            worksheet.cell(row=i, column=3, value=float(data_k_volume))

            workbook.save(file_path)


class Contract:
    def __init__(self, initial_balance=10000):
        self.balance = initial_balance
        self.grid_size = 1000
        self.flag = "INIT"
        self.amount = 0

    def execute_trade(self, action, quantity, price):
        cost = quantity * price
        if action == "buy" and cost <= self.balance:
            # 调用买入合约接口
            self.balance -= cost
            logging.info(f"Bought {quantity} contracts at price {price} for ${cost}")
        elif action == "sell":
            # 调用卖出合约接口
            self.balance += cost
            logging.info(f"Sold {quantity} contracts at price {price} for ${cost}")
        else:
            logging.info("Insufficient balance to buy the contract.")

    def apply_grid_trading(self, current_price):
        lower_price = int(current_price / self.grid_size) * self.grid_size
        upper_price = lower_price + self.grid_size

        self.buy_contract(1, lower_price)
        self.sell_contract(1, upper_price)

    def get_balance(self, price):
        # 调用查询余额和coin数量的接口
        logging.info(f"Current balance: ${self.balance}")
        logging.info(f"Current amount: {self.amount}")
        logging.info(f"Current asset: {self.amount * price + self.balance}")

    def maximize_profit(self, prediction, price):

        if prediction > price:
            # 如果预测价格为上涨，则买入期货合约
            if self.amount == 0:
                self.amount = round(float(self.balance) / price, 2)
            else:
                logging.info(f"之前已买入，现在趋势是继续上涨，继续观望。")
            self.execute_trade("buy", self.amount, price)
        elif prediction < price:
            # 如果预测价格为下跌，则卖出期货合约
            if self.amount > 0:
                self.execute_trade("sell", self.amount, price)
                self.amount = 0
            else:
                logging.info(f"现在还没买入，但是趋势是下跌，继续等候上涨信号再买入。")
        else:
            logging.info(f"价格不变，继续观望。")


        self.get_balance(price)  # 查询当前余额


if __name__ == "__main__":
    # 加载数据
    file_path = 'ethcoin_data_with_indicators.xlsx'  # 更改为您的文件路径
    output_file_path = "compare_eth_data.txt"
    logging.basicConfig(filename='eth_execute.log', level=logging.INFO)

    AI_Trainer = AIPredict()
    Gateioget = GateIO_Api()
    Contract = Contract(initial_balance=100)

    logging.info(f"===========initialize===========")
    minutes_to_add = 5

    while True:
        # 清理数据
        AI_Trainer.clean_data(file_path)
        logging.info(f"===========clean data ok===========")
        # get得到数据
        Gateioget.get_current_data_api(file_path)
        logging.info(f"===========get new 1000 datas ok===========")
        # 按时间升序进行排序
        AI_Trainer.sort_date_value(file_path)
        logging.info(f"===========sort datas ok===========")
        # 计算指数（去掉空值）
        AI_Trainer.caculate_data(file_path)  # 计算均值等技术指标
        logging.info(f"===========calulate ma rsi upper ok===========")
        # 训练数据
        bitcoin_data = pd.read_excel(file_path)
        model = AI_Trainer.train_data(bitcoin_data)  # 训练模型
        logging.info(f"===========train data ok===========")
        # 预测5min之后的数据
        furture_price = AI_Trainer.predict_next(model, bitcoin_data)
        logging.info(f"Next Furture Price: ${furture_price}")
        # 得到当前的价格
        current_price = Gateioget.get_current_price()
        logging.info(f"Current Price: ${current_price}")
        # 执行交易
        Contract.maximize_profit(float(furture_price), float(current_price))
        logging.info(f"===========contract business ok===========")
        # 等待5min
        Gateioget.get_right_time(minutes_to_add)
        # 得到5min后的真实价格
        true_price = Gateioget.get_current_price()
        logging.info(f"True Price after ${minutes_to_add}min: ${true_price}")
        with open(output_file_path, mode='a') as file:
            # 将时间、价格和预测值以制表符分隔的格式写入文件
            data_k_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            file.write(f"{current_price}\t{furture_price}\t{data_k_date}\t{true_price}\n")
        logging.info(f"===========record data ok===================================================")
