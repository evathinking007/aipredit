import warnings
import openpyxl
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
from datetime import timedelta
from datetime import datetime
import time
import requests
import logging


class AIPredict:
    def __init__(self):
        pass

    def create_excel(self, file_path):
        # 创建一个新的工作簿
        wb = openpyxl.Workbook()

        # 获取默认的工作表
        sheet = wb.active
        sheet['A1'] = 'Date'
        sheet['B1'] = 'Price'
        sheet['C1'] = 'Volume'
        sheet['D1'] = 'MA'
        sheet['E1'] = 'RSI'
        sheet['F1'] = 'Upper_BB'
        sheet['G1'] = 'Lower_BB'
        sheet['H1'] = 'Future_Price'
        wb.save(file_path)

    def caculate_data(self, file_path):
        df = pd.read_excel(file_path)
        ma_period = 20  # 移动平均的周期
        df['MA'] = df['Price'].rolling(window=ma_period).mean()

        # 计算相对强弱指标（RSI）
        rsi_period = 14  # RSI的周期
        price_diff = df['Price'].diff()
        gain = price_diff.where(price_diff > 0, 0)
        loss = -price_diff.where(price_diff < 0, 0)
        avg_gain = gain.rolling(window=rsi_period).mean()
        avg_loss = loss.rolling(window=rsi_period).mean()
        rs = avg_gain / avg_loss
        rsi = 100 - (100 / (1 + rs))
        df['RSI'] = rsi

        # 计算布林带的上限（Upper_BB）和下限（Lower_BB）
        bb_period = 20  # 布林带的周期
        std_dev = df['Price'].rolling(window=bb_period).std()
        df['Upper_BB'] = df['MA'] + (2 * std_dev)
        df['Lower_BB'] = df['MA'] - (2 * std_dev)
        df['Future_Price'] = None

        # 回写到Excel文件
        df.to_excel(file_path, index=False)

    def sort_date_value(self, file_path):
        bitcoin_data = pd.read_excel(file_path)
        bitcoin_data.sort_values(by='Date', ascending=True)
        bitcoin_data.to_excel(file_path, index=False)

    def trading_strategy(self, file_path, current_price):
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个表格（如果有多个表格，请根据实际情况进行选择）
        sheet = workbook.active
        last_row = sheet.max_row
        lower_bband = sheet.cell(row=last_row, column='Lower_BB').value
        upper_bband = sheet.cell(row=last_row, column='Upper_BB').value
        rsi = sheet.cell(row=last_row, column='RSI').value
        rsi_buy_threshold = 30  # 买入阈值
        rsi_sell_threshold = 70  # 卖出阈值
        mean_reversion_threshold = sheet.cell(row=last_row, column='MA').value

        if current_price < lower_bband and rsi < rsi_buy_threshold:
            signal = 'BUY'
        elif current_price > upper_bband and rsi > rsi_sell_threshold:
            signal = 'SELL'
        elif current_price > mean_reversion_threshold and rsi > rsi_sell_threshold:
            signal = 'SELL'
        elif current_price < mean_reversion_threshold and rsi < rsi_buy_threshold:
            signal = 'BUY'
        else:
            signal = 'HOLD'

        return signal


class GateIO_Api:
    def __init__(self):
        pass

    def get_right_time(self, minutes_to_add):
        # 获取当前时间
        current_time = datetime.now()
        # 计算下一个整点时刻
        next_time = current_time + timedelta(minutes=minutes_to_add)
        next_time = next_time.replace(second=0)

        # 计算等待时间，直到下一个整点时刻
        time_to_wait = (next_time - current_time).total_seconds()
        logging.info(f"等待到下一个整点时刻（{next_time.strftime('%d/%m/%Y, %H:%M:%S')}）")
        time.sleep(time_to_wait)

    def get_current_price(self, file_path, coin_name):
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        worksheet = workbook.active
        host = "https://api.gateio.ws"
        prefix = "/api/v4"
        headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}

        url = '/spot/candlesticks'
        query_param = 'currency_pair=' + coin_name + '_USDT&limit=1'
        r = requests.request('GET', host + prefix + url + "?" + query_param, headers=headers)
        data_k_date = datetime.fromtimestamp(int(r.json()[0][0]))
        data_k_volume = r.json()[0][6]
        data_k_price = r.json()[0][2]
        max_row = worksheet.max_row
        worksheet.cell(row=max_row, column=1, value=data_k_date)
        worksheet.cell(row=max_row, column=2, value=float(data_k_price))
        worksheet.cell(row=max_row, column=3, value=float(data_k_volume))

        workbook.save(file_path)
        return data_k_price

    def get_current_data_api(self, file_path, coin_name):
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        worksheet = workbook.active

        host = "https://api.gateio.ws"
        prefix = "/api/v4"
        headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}

        url = '/spot/candlesticks'
        query_param = 'currency_pair=' + coin_name + '_USDT&limit=1000&interval=5m'

        r = requests.request('GET', host + prefix + url + "?" + query_param, headers=headers)
        data_rows_num = len(r.json())
        logging.info("===========get " + str(data_rows_num) + " k datas===========")
        worksheet.insert_rows(data_rows_num)
        for i in range(2, data_rows_num):
            data_k_date = datetime.fromtimestamp(int(r.json()[i][0]))
            data_k_price = r.json()[i][2]
            data_k_volume = r.json()[i][6]

            worksheet.cell(row=i, column=1, value=data_k_date)
            worksheet.cell(row=i, column=2, value=float(data_k_price))
            worksheet.cell(row=i, column=3, value=float(data_k_volume))

            workbook.save(file_path)


class Contract:
    def __init__(self, initial_balance=10000):
        self.balance = initial_balance
        self.amount = 0

    def execute_trade(self, action, quantity, price):
        cost = quantity * price
        if action == "buy" and cost <= self.balance:
            # 调用买入合约接口
            self.balance -= cost
            logging.info(f"Bought {quantity} contracts at price {price} for ${cost}")
        elif action == "sell":
            # 调用卖出合约接口
            self.balance += cost
            logging.info(f"Sold {quantity} contracts at price {price} for ${cost}")
        else:
            logging.info("Insufficient balance to buy the contract.")

    def get_balance(self, price):
        # 调用查询余额和coin数量的接口
        logging.info(f"Current balance: ${self.balance}")
        logging.info(f"Current amount: {self.amount}")
        logging.info(f"Current asset: {self.amount * price + self.balance}")

    def maximize_profit(self, prediction, price):

        if prediction == "BUY":
            # 如果预测价格为上涨，则买入期货合约
            if self.amount == 0:
                self.amount = round(float(self.balance) / price, 2)
            self.execute_trade("buy", self.amount, price)
        elif prediction == "SELL":
            # 如果预测价格为下跌，则卖出期货合约
            if self.amount > 0:
                self.execute_trade("sell", self.amount, price)
                self.amount = 0
            else:
                logging.info(f"现在还没买入")
        else:
            logging.info(f"价格不变，继续观望。")

        self.get_balance(price)  # 查询当前余额


if __name__ == "__main__":
    # 加载数据
    coin_name = "ETH"
    now_time = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
    file_path = coin_name + "_" + now_time + ".xlsx"  # 更改为您的文件路径
    logging.basicConfig(filename=coin_name + '_execute.log', level=logging.INFO)

    AI_Trainer = AIPredict()
    Gateioget = GateIO_Api()
    Contract = Contract(initial_balance=100)
    minutes_to_add = 1

    logging.info(f"===========initialize==============================================")
    # 创建excel
    AI_Trainer.create_excel(file_path)
    logging.info(f"===========clean data ok===========")
    # 读取5min数据
    Gateioget.get_current_data_api(file_path, coin_name)
    logging.info(f"===========get new 1000 datas ok===========")
    # 按时间升序进行排序
    AI_Trainer.sort_date_value(file_path)
    logging.info(f"===========sort datas ok===========")

    while True:
        # 获取当前值
        current_price = Gateioget.get_current_price(file_path, coin_name)
        logging.info(f"Current Price: ${current_price}")

        # 计算指数
        AI_Trainer.caculate_data(file_path)  # 计算均值等技术指标
        logging.info(f"===========calulate ma rsi upper ok===========")

        # 判断信号
        signal = AI_Trainer.trading_strategy(file_path, current_price)
        logging.info(f"===========signal is {signal}===========")

        # 执行交易
        Contract.maximize_profit(signal, float(current_price))
        logging.info(f"===========contract business ok===========")

        # 等待5min
        Gateioget.get_right_time(minutes_to_add)
