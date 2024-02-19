import warnings
import openpyxl
from datetime import timedelta
from datetime import datetime, timezone
import time
import requests
import logging
import threading
import math
import pandas as pd
from gate_api import ApiClient, Configuration, FuturesApi, FuturesOrder, Transfer, WalletApi
from gate_api.exceptions import GateApiException
from decimal import Decimal as D, ROUND_UP, getcontext


def setup_logger(name, log_file, level=logging.INFO):
    # """Function to setup as many loggers as you want"""
    handler = logging.FileHandler(log_file)
    formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
    handler.setFormatter(formatter)

    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.addHandler(handler)

    return logger

    # 为当前线程设置独立的日志记录器


class GateIO_Api:
    def __init__(self, logger):
        self.logger = logger
        pass

    def get_right_time(self, minutes_to_add):
        # 获取当前时间
        current_time = datetime.now()
        # 计算下一个整点时刻
        next_time = current_time + timedelta(minutes=minutes_to_add)
        next_time = next_time.replace(second=0)

        # 计算等待时间，直到下一个整点时刻
        time_to_wait = (next_time - current_time).total_seconds()
        self.logger.info(f"等待到下一个整点时刻（{next_time.strftime('%d/%m/%Y, %H:%M:%S')}）")
        time.sleep(time_to_wait)

    def create_excel(self, file_path):
        # 创建一个新的工作簿
        wb = openpyxl.Workbook()

        # 获取默认的工作表
        sheet = wb.active
        sheet['A1'] = 'buy_time'
        sheet['B1'] = 'buy_price'
        sheet['C1'] = 'buy_amount'
        sheet['D1'] = 'buy_percentage'
        sheet['E1'] = 'sell_time'
        sheet['F1'] = 'sell_price'
        sheet['G1'] = 'sell_amount'
        sheet['H1'] = 'sell_percentage'
        sheet['I1'] = 'direction'
        wb.save(file_path)

    def caculate_zhibiao(self, history_file_path):
        df = pd.read_excel(history_file_path)
        ma_period = 20  # 移动平均的周期
        df['MA'] = df['Close'].rolling(window=ma_period).mean()

        # 计算相对强弱指标（RSI）
        rsi_period = 14  # RSI的周期
        price_diff = df['Close'].diff()
        gain = price_diff.where(price_diff > 0, 0)
        loss = -price_diff.where(price_diff < 0, 0)
        avg_gain = gain.rolling(window=rsi_period).mean()
        avg_loss = loss.rolling(window=rsi_period).mean()
        rs = avg_gain / avg_loss
        rsi = 100 - (100 / (1 + rs))
        df['RSI'] = rsi

        # 计算布林带的上限（Upper_BB）和下限（Lower_BB）
        bb_period = 20  # 布林带的周期
        std_dev = df['Close'].rolling(window=bb_period).std()
        df['Upper_BB'] = df['MA'] + (2 * std_dev)
        df['Lower_BB'] = df['MA'] - (2 * std_dev)

        # 计算ATR真实波动曲线
        atr_period = 14
        df['H-L'] = df['High'] - df['Low']
        df['H-PC'] = abs(df['High'] - df['Close'].shift(1))
        df['L-PC'] = abs(df['Low'] - df['Close'].shift(1))
        df['TR'] = df[['H-L', 'H-PC', 'L-PC']].max(axis=1)
        df['ATR'] = df['TR'].rolling(window=atr_period).mean()

        # df = df.iloc[20:]

        # 回写到Excel文件
        df.to_excel(history_file_path, index=False)

    def get_current_price(self, file_path, coin_name):
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        worksheet = workbook.active
        host = "https://fx-api-testnet.gateio.ws"
        prefix = "/api/v4"
        headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}
        url = '/futures/usdt/candlesticks'
        query_param = 'contract=' + coin_name + '_USDT&limit=1'
        r = requests.request('GET', host + prefix + url + "?" + query_param, headers=headers)

        data_k_date = datetime.fromtimestamp(int(r.json()[0]['t']))
        data_k_close = r.json()[0]['c']
        data_k_high = r.json()[0]['h']
        data_k_low = r.json()[0]['l']
        data_k_open = r.json()[0]['o']

        max_row = worksheet.max_row
        worksheet.cell(row=max_row, column=1, value=data_k_date)
        worksheet.cell(row=max_row, column=2, value=float(data_k_open))
        worksheet.cell(row=max_row, column=3, value=float(data_k_high))
        worksheet.cell(row=max_row, column=4, value=float(data_k_low))
        worksheet.cell(row=max_row, column=5, value=float(data_k_close))
        workbook.save(file_path)

        return data_k_close

    def get_1day_base_price(self, file_path, coin_name):
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        worksheet = workbook.active

        host = "https://api.gateio.ws"
        prefix = "/api/v4"
        headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}

        url = '/spot/candlesticks'
        query_param = 'currency_pair=' + coin_name + '_USDT&limit=1&interval=1d'

        r = requests.request('GET', host + prefix + url + "?" + query_param, headers=headers)
        print(r.json()[0])
        return r.json()[0][3]

    def save_contract_duo_buy_trade_into_excel(self, file_path, trade):
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        worksheet = workbook.active
        max_row = worksheet.max_row
        worksheet.cell(row=max_row + 1, column=1, value=trade["buy_time"])
        worksheet.cell(row=max_row + 1, column=2, value=trade["buy_price"])
        worksheet.cell(row=max_row + 1, column=3, value=trade["buy_amount"])
        worksheet.cell(row=max_row + 1, column=4, value=trade["buy_percentage"])
        worksheet.cell(row=max_row + 1, column=5, value=trade["sell_time"])
        worksheet.cell(row=max_row + 1, column=6, value=trade["sell_price"])
        worksheet.cell(row=max_row + 1, column=7, value=trade["sell_amount"])
        worksheet.cell(row=max_row + 1, column=8, value=trade["sell_percentage"])
        worksheet.cell(row=max_row + 1, column=9, value=trade["direction"])

        workbook.save(file_path)

    def save_contract_duo_sell_trade_into_excel(self, file_path, trade):
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        worksheet = workbook.active
        max_row = worksheet.max_row
        for i in range(2, max_row + 1):
            if worksheet.cell(row=i, column=1).value == trade["buy_time"]:
                # self.logger.info("卖出写入判断成功")
                worksheet.cell(row=i, column=5, value=trade["sell_time"])
                worksheet.cell(row=i, column=6, value=trade["sell_price"])
                worksheet.cell(row=i, column=7, value=trade["sell_amount"])
                worksheet.cell(row=i, column=8, value=trade["sell_percentage"])
        workbook.save(file_path)

    def save_contract_kong_sell_trade_into_excel(self, file_path, trade):
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        worksheet = workbook.active
        max_row = worksheet.max_row
        worksheet.cell(row=max_row + 1, column=1, value=trade["buy_time"])
        worksheet.cell(row=max_row + 1, column=2, value=trade["buy_price"])
        worksheet.cell(row=max_row + 1, column=3, value=trade["buy_amount"])
        worksheet.cell(row=max_row + 1, column=4, value=trade["buy_percentage"])
        worksheet.cell(row=max_row + 1, column=5, value=trade["sell_time"])
        worksheet.cell(row=max_row + 1, column=6, value=trade["sell_price"])
        worksheet.cell(row=max_row + 1, column=7, value=trade["sell_amount"])
        worksheet.cell(row=max_row + 1, column=8, value=trade["sell_percentage"])
        worksheet.cell(row=max_row + 1, column=9, value=trade["direction"])
        workbook.save(file_path)

    def save_contract_kong_buy_trade_into_excel(self, file_path, trade):
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        worksheet = workbook.active
        max_row = worksheet.max_row
        for i in range(2, max_row + 1):
            if worksheet.cell(row=i, column=5).value == trade["sell_time"]:
                worksheet.cell(row=i, column=1, value=trade["buy_time"])
                worksheet.cell(row=i, column=2, value=trade["buy_price"])
                worksheet.cell(row=i, column=3, value=trade["buy_amount"])
                worksheet.cell(row=i, column=4, value=trade["buy_percentage"])
        workbook.save(file_path)

    def round_down(self, value, decimals):
        factor = 10 ** decimals
        return math.floor(value * factor) / factor

    def wait_to_sell_or_buy(self, file_path):
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        worksheet = workbook.active
        max_row = worksheet.max_row
        wait_to_sell = []
        wait_to_buy = []
        for i in range(1, max_row + 1):
            if worksheet.cell(row=i, column=1).value is not None and worksheet.cell(row=i, column=5).value is None:
                sell_trade = {
                    'buy_time': worksheet.cell(row=i, column=1).value,
                    'buy_price': worksheet.cell(row=i, column=2).value,
                    'buy_amount': worksheet.cell(row=i, column=3).value,
                    'buy_percentage': worksheet.cell(row=i, column=4).value,
                    'sell_time': worksheet.cell(row=i, column=5).value,
                    'sell_price': worksheet.cell(row=i, column=6).value,
                    'sell_amount': worksheet.cell(row=i, column=7).value,
                    'sell_percentage': worksheet.cell(row=i, column=8).value,
                    'direction': worksheet.cell(row=i, column=9).value,
                }
                wait_to_sell.append(sell_trade)
        for i in range(1, max_row + 1):
            if worksheet.cell(row=i, column=1).value is None and worksheet.cell(row=i, column=5).value is not None:
                buy_trade = {
                    'buy_time': worksheet.cell(row=i, column=1).value,
                    'buy_price': worksheet.cell(row=i, column=2).value,
                    'buy_amount': worksheet.cell(row=i, column=3).value,
                    'buy_percentage': worksheet.cell(row=i, column=4).value,
                    'sell_time': worksheet.cell(row=i, column=5).value,
                    'sell_price': worksheet.cell(row=i, column=6).value,
                    'sell_amount': worksheet.cell(row=i, column=7).value,
                    'sell_percentage': worksheet.cell(row=i, column=8).value,
                    'direction': worksheet.cell(row=i, column=9).value,
                }
                wait_to_buy.append(buy_trade)
        return wait_to_sell, wait_to_buy

    def get_history_data(self, coin_name, start_time, period):
        wb = openpyxl.Workbook()

        # 获取默认的工作表
        sheet = wb.active
        sheet['A1'] = 'Date'
        sheet['B1'] = 'Open'
        sheet['C1'] = 'High'
        sheet['D1'] = 'Low'
        sheet['E1'] = 'Close'
        file_path = coin_name + "_" + str(start_time) + "_history_data_" + period + ".xlsx"
        wb.save(file_path)

        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        worksheet = workbook.active

        host = "https://fx-api-testnet.gateio.ws"
        prefix = "/api/v4"
        headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}
        url = '/futures/usdt/candlesticks'
        query_param = 'contract=' + coin_name + '_USDT&limit=2000&interval=' + period

        r = requests.request('GET', host + prefix + url + "?" + query_param, headers=headers)
        data_rows_num = len(r.json())
        self.logger.info("===========geting " + str(data_rows_num) + " k datas...===========")
        worksheet.insert_rows(data_rows_num)
        for i in range(2, data_rows_num):
            data_k_date = datetime.fromtimestamp(int(r.json()[i]['t']))
            data_k_close = r.json()[i]['c']
            data_k_high = r.json()[i]['h']
            data_k_low = r.json()[i]['l']
            data_k_open = r.json()[i]['o']

            worksheet.cell(row=i, column=1, value=data_k_date)
            worksheet.cell(row=i, column=2, value=float(data_k_open))
            worksheet.cell(row=i, column=3, value=float(data_k_high))
            worksheet.cell(row=i, column=4, value=float(data_k_low))
            worksheet.cell(row=i, column=5, value=float(data_k_close))

            workbook.save(file_path)
        return file_path


class Contract:
    def __init__(self, logger, coin_name, initial_balance=10000):
        self.logger = logger
        self.settle = "usdt"
        self.api_key = "435ab3bbd208058288eba0e341229df3"
        self.api_secret = "c79daea180bca489ad673b02a1b1679ca0cd9ae8fb4b86c3741cd4a8116606f0"
        self.host_used = "https://fx-api-testnet.gateio.ws/api/v4"
        self.contract = coin_name + '_USDT'
        self.balance = self.get_balance()

    def submit_order(self, futures_api, order_size, price):
        order = FuturesOrder(contract=self.contract, size=order_size, price=price, tif='gtc')
        try:
            order_response = futures_api.create_futures_order(self.settle, order)
        except GateApiException as ex:
            self.logger.error("error encountered creating futures order: %s", ex)
            return
        self.logger.info("order %s created with status: %s", order_response.id, order_response.status)

        if order_response.status == 'open':
            futures_order = futures_api.get_futures_order(self.settle, str(order_response.id))
            self.logger.info("order %s status %s, total size %s, left %s", futures_order.id, futures_order.status,
                             futures_order.size, futures_order.left)
            # futures_api.cancel_futures_order(self.settle, str(futures_order.id))
            # logger.info("order %s cancelled", futures_order.id)
        else:
            time.sleep(0.2)
            order_trades = futures_api.get_my_trades(self.settle, contract=self.contract, order=order_response.id)
            assert len(order_trades) > 0
            trade_size = 0
            for t in order_trades:
                assert t.order_id == str(order_response.id)
                trade_size += t.size
                self.logger.info("order %s filled size %s with price %s", t.order_id, t.size, t.price)
            assert trade_size == order_size

    def get_order_size(self, futures_api, quantity, leverage):
        futures_contract = futures_api.get_futures_contract(self.settle, self.contract)
        assert futures_contract.quanto_multiplier
        order_size = D(quantity) * D(leverage) / D(futures_contract.quanto_multiplier)
        return int(order_size)

    def execute_trade(self, action, quantity, price):
        cost = quantity * price
        print(f"{action}, amount: {quantity} contracts at price {price} for ${cost}")

        config = Configuration(key=self.api_key, secret=self.api_secret, host=self.host_used)
        futures_api = FuturesApi(ApiClient(config))

        # 设置杠杆
        leverage = "50"
        futures_api.update_position_leverage(self.settle, self.contract, leverage)

        # 开始交易
        self.balance = self.get_balance()  # 查询当前余额

        if action == "buy" and cost <= self.balance:
            # 调用买入合约接口
            order_size = self.get_order_size(futures_api, quantity, leverage)
            self.submit_order(futures_api, order_size, price)
            self.logger.info(f"Bought {quantity} contracts at price {price} for ${cost}")

        elif action == "sell":
            order_size = self.get_order_size(futures_api, quantity)
            self.submit_order(futures_api, -order_size, price)
            self.logger.info(f"Sold {quantity} contracts at price {price} for ${cost}")
        else:
            self.logger.info("参数错误")

        self.get_balance()  # 查询当前余额

    def get_balance(self):
        # 调用查询余额和coin数量的接口
        config = Configuration(key=self.api_key, secret=self.api_secret, host=self.host_used)
        futures_api = FuturesApi(ApiClient(config))
        futures_account = futures_api.list_futures_accounts(self.settle)
        self.balance = float(futures_account.total)
        self.logger.info(f"Current balance: ${self.balance}")
        return self.balance


def trading_run(coin_name, period, initial_balance=1000, investment_per_trade=100, grid_start=1, grid_step=0.5,
                sell_step=1.025, buy_step=0.975):
    now_time = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
    minutes_to_add = 5

    # 创建日志把手
    log_file_name = f"{coin_name}_{now_time}_execute.log"
    logger = setup_logger(coin_name, log_file_name)

    # 初始化交易类
    gateioget = GateIO_Api(logger)
    contract = Contract(logger, coin_name, initial_balance)

    # 策略开始运行
    logger.info("策略开始运行")

    # 获取账户信息
    contract.get_balance()

    # 创建交易记录文件
    file_path = coin_name + "_trade.xlsx"  # 更改为您的文件路径
    gateioget.create_excel(file_path)

    # 得到最近2000条历史数据
    # history_file_path = gateioget.get_history_data(coin_name, now_time, period)
    history_file_path = 'ETH_2024-02-18-22-12-52_history_data_5m.xlsx'
    while True:
        # 得到当前价格
        now_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        current_price = float(gateioget.get_current_price(history_file_path, coin_name))
        logger.info("当前的价格是：" + str(current_price) + ",时间是:" + str(now_time))

        # 开始计算指标（ma）
        gateioget.caculate_zhibiao(history_file_path)

        # 选择第一个工作表
        workbook = openpyxl.load_workbook(history_file_path)
        worksheet = workbook.active
        max_row = worksheet.max_row

        # 得到网格的基线价格，即ma值
        base_price = float(worksheet.cell(row=max_row, column=6).value)
        logger.info("当前价格的均线值是：" + str(base_price))

        # 开始运行网格下行
        sell_trades, buy_trades = gateioget.wait_to_sell_or_buy(file_path)
        if len(sell_trades) < 5:
            investment_per_trade2 = investment_per_trade
            # 得到均线偏离值
            buy_percentage = round((1 - current_price / base_price) * 100, 2)
            logger.info("当前的价格偏离均线：-" + str(buy_percentage) + "%")
            if buy_percentage >= grid_start + 4 * grid_step:
                signal="buy"
                # investment_per_trade2 = investment_per_trade * 3
            elif buy_percentage >= grid_start + 3 * grid_step:
                signal="sell"
                # investment_per_trade2 = investment_per_trade * 2.5
            elif buy_percentage >= grid_start + 2 * grid_step:
                signal="sell"
                # investment_per_trade2 = investment_per_trade * 2
            elif buy_percentage >= grid_start + 1 * grid_step:
                signal="sell"
                # investment_per_trade2 = investment_per_trade * 1.5
            elif buy_percentage >= grid_start:
                signal="sell"
                # investment_per_trade2 = investment_per_trade * 1
            else:
                signal="hold"
                # investment_per_trade2 = 0
        if len(buy_trades) < 5:
            investment_per_trade2 = investment_per_trade
            sell_percentage = round((current_price / base_price - 1) * 100, 2)
            logger.info("当前的价格偏离均线：" + str(sell_percentage) + "%")
            if sell_percentage >= grid_start + 4 * grid_step:
                signal2="sell"
                # investment_per_trade2 = investment_per_trade * 3
            elif sell_percentage >= grid_start + 3 * grid_step:
                signal2="buy"
                # investment_per_trade2 = investment_per_trade * 2.5
            elif sell_percentage >= grid_start + 2 * grid_step:
                signal2="buy"
                # investment_per_trade2 = investment_per_trade * 2
            elif sell_percentage >= grid_start + 1 * grid_step:
                signal2="buy"
                # investment_per_trade2 = investment_per_trade * 1.5
            elif sell_percentage >= grid_start:
                signal2="buy"
                # investment_per_trade2 = investment_per_trade * 1
            else:
                signal2="hold"
                # investment_per_trade2 = 0    
            
        sell_trades, buy_trades = gateioget.wait_to_sell_or_buy(file_path)   
        if signal =="sell" or signal2=="sell":
            sell_amount = gateioget.round_down(investment_per_trade2 / current_price, 2)  # 每次买入investment_per_trade
            if sell_amount > 0:
                sell_price = current_price
                sell_time = now_time
                contract.execute_trade("sell", sell_amount, sell_price)
                logger.info("===========做空==================")

                # 添加到交易记录
                sell_trade = {
                    'buy_time': None,
                    'buy_price': None,
                    'buy_amount': None,
                    'buy_percentage': None,
                    'sell_time': sell_time,
                    'sell_price': sell_price,
                    'sell_amount': sell_amount,
                    'sell_percentage': sell_percentage,
                    'direction': '空',
                }
                # 这里都用一个方法
                gateioget.save_contract_kong_sell_trade_into_excel(file_path, sell_trade)
        if signal == "buy" or signal2=="buy":
                buy_amount = gateioget.round_down(investment_per_trade2 / current_price, 2)  # 每次买入investment_per_trade
                if buy_amount > 0:
                    buy_price = current_price
                    buy_time = now_time
                    contract.execute_trade("buy", buy_amount, buy_price)
                    logger.info("===========做多==================")

                    # 添加到交易记录
                    buy_trade = {
                        'buy_time': buy_time,
                        'buy_price': buy_price,
                        'buy_amount': buy_amount,
                        'buy_percentage': buy_percentage,
                        'sell_time': None,
                        'sell_price': None,
                        'sell_amount': None,
                        'sell_percentage': None,
                        'direction': '多',
                    }
                    gateioget.save_contract_duo_buy_trade_into_excel(file_path, buy_trade)
        
       
                buy_amount = gateioget.round_down(investment_per_trade2 / current_price, 2)  # 每次买入investment_per_trade
                if buy_amount > 0:
                    buy_price = current_price
                    buy_time = now_time
                    contract.execute_trade("buy", buy_amount, buy_price)
                    logger.info("===========做多==================")

                    # 添加到交易记录
                    buy_trade = {
                        'buy_time': buy_time,
                        'buy_price': buy_price,
                        'buy_amount': buy_amount,
                        'buy_percentage': buy_percentage,
                        'sell_time': None,
                        'sell_price': None,
                        'sell_amount': None,
                        'sell_percentage': None,
                        'direction': '多',
                    }
                    gateioget.save_contract_duo_buy_trade_into_excel(file_path, buy_trade)
        
        # 开始执行结对卖出做多的合约
        sell_trades, buy_trades = gateioget.wait_to_sell_or_buy(file_path)
        for trade in sell_trades:
            sell_target_price = float(trade['buy_price']) * sell_step  # 目标卖出价格
            if current_price >= sell_target_price:
                if current_price > base_price * 1.01:
                    logger.info("当前的卖出合约价格在均线之上1%，建议继续持有。")
                    continue
                else:
                    sell_amount = trade['buy_amount']
                    sell_price = current_price
                    sell_time = now_time
                    sell_percentage = round((sell_price / float(trade['buy_price']) - 1) * 100, 2)
                    contract.execute_trade("sell", sell_amount, sell_price)
                    logger.info("===========做多卖出==================")
                    # 更新交易记录
                    sell_trade = {
                        'buy_time': trade['buy_time'],
                        'buy_price': trade['buy_price'],
                        'buy_amount': trade['buy_amount'],
                        'buy_percentage': trade['buy_percentage'],
                        'sell_time': sell_time,
                        'sell_price': sell_price,
                        'sell_amount': sell_amount,
                        'sell_percentage': sell_percentage,
                        'direction': trade['direction'],
                    }
                    gateioget.save_contract_duo_sell_trade_into_excel(file_path, sell_trade)

     
            investment_per_trade2 = investment_per_trade
            sell_percentage = round((current_price / base_price - 1) * 100, 2)
            logger.info("当前的价格偏离均线：" + str(sell_percentage) + "%")
            if sell_percentage >= grid_start + 4 * grid_step:
                investment_per_trade2 = investment_per_trade * 3
            elif sell_percentage >= grid_start + 3 * grid_step:
                investment_per_trade2 = investment_per_trade * 2.5
            elif sell_percentage >= grid_start + 2 * grid_step:
                investment_per_trade2 = investment_per_trade * 2
            elif sell_percentage >= grid_start + 1 * grid_step:
                investment_per_trade2 = investment_per_trade * 1.5
            elif sell_percentage >= grid_start:
                investment_per_trade2 = investment_per_trade * 1
            else:
                investment_per_trade2 = 0

            if investment_per_trade2 > 0:
                sell_amount = gateioget.round_down(investment_per_trade2 / current_price, 2)  # 每次买入investment_per_trade
                if sell_amount > 0:
                    sell_price = current_price
                    sell_time = now_time
                    contract.execute_trade("sell", sell_amount, sell_price)
                    logger.info("===========做空==================")

                    # 添加到交易记录
                    sell_trade = {
                        'buy_time': None,
                        'buy_price': None,
                        'buy_amount': None,
                        'buy_percentage': None,
                        'sell_time': sell_time,
                        'sell_price': sell_price,
                        'sell_amount': sell_amount,
                        'sell_percentage': sell_percentage,
                        'direction': '空',
                    }
                    # 这里都用一个方法
                    gateioget.save_contract_kong_sell_trade_into_excel(file_path, sell_trade)

        # 开始执行结对买入做空的合约
        sell_trades, buy_trades = gateioget.wait_to_sell_or_buy(file_path)
        for trade in buy_trades:
            buy_target_price = float(trade['sell_price']) * buy_step  # 目标买入价格
            if current_price <= buy_target_price:
                if current_price < base_price * 0.99:
                    logger.info("当前的平仓合约价格在均线之下1%，建议继续持有。")
                    continue
                else:
                    buy_amount = trade['sell_amount']
                    buy_price = current_price
                    buy_time = now_time
                    buy_percentage = round((float(trade['sell_price']) / buy_price - 1) * 100, 2)
                    contract.execute_trade("buy", buy_amount, buy_price)
                    logger.info("===========做空买入==================")

                    # 更新交易记录
                    buy_trade = {
                        'buy_time': buy_time,
                        'buy_price': buy_price,
                        'buy_amount': buy_amount,
                        'buy_percentage': buy_percentage,
                        'sell_time': trade['sell_time'],
                        'sell_price': trade['buy_price'],
                        'sell_amount': trade['sell_amount'],
                        'sell_percentage': trade['sell_percentage'],
                        'direction': trade['direction'],
                    }
                    gateioget.save_contract_kong_buy_trade_into_excel(file_path, buy_trade)

                    # 间隔时间后再读取下一次数据

        gateioget.get_right_time(minutes_to_add)

    # # 这个是跑回测的


# 这个是跑实盘的
if __name__ == "__main__":
    # 加载数据
    coins = [
        {'name': 'ETH', 'init': 1000, 'k_period': '5m', 'invest_per_trade': 100, 'grid_start': 0, 'grid_step': 0,
         'sell_step': 1, 'buy_step': 1}
        # {'name':'BTC','init':10000,'k_period':'5m','invest_per_trade':2000,'grid_start':0.5,'grid_step':0.2,'sell_step':1.03,'buy_step':0.994},
        # {'name':'BNB','init':1000,'k_period':'5m','invest_per_trade':100,'grid_start':0.5,'grid_step':0.2,'sell_step':1.03,'buy_step':0.994}
    ]

    threads = []
    for coin in coins:
        thread = threading.Thread(target=trading_run, args=(
            coin['name'], coin['k_period'], coin['init'], coin['invest_per_trade'], coin['grid_start'],
            coin['grid_step'],
            coin['sell_step'], coin['buy_step']))
        threads.append(thread)

    # 启动所有线程
    for thread in threads:
        thread.start()

    # 等待所有线程完成
    for thread in threads:
        thread.join()
